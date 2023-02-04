from log.logger import log
from openpyxl import load_workbook
import datetime
import os.path
import cx_Oracle


OS = 'WINDOWS'
UPLOAD_PATH = '04.02.2023'
LOAD_TEST_RU = True
LOAD_TEST_KZ = True
LOAD_TEST_EN = True
IMAGE_VERSION = '04022023'

if OS != 'WINDOWS':
    LIB_DIR = r'/home/pdd/instantclient_21_4'
else:
    LIB_DIR = r'd:/install/oracle/instantclient_21_3'


cx_Oracle.init_oracle_client(lib_dir=LIB_DIR)

def load_partition_questions(id_task, lang, partition_number, subpartition_number, file_name):
    theme_number = 2
    s_now = datetime.datetime.now()
    partition_messages = []
    if OS == 'unix':
       file_path = UPLOAD_PATH + '/' + lang + '/' + file_name
    else:
       file_path = UPLOAD_PATH + '\\' + lang + '\\' + file_name

    #
    path = os.path.normpath(file_path)

    log.info("Download started: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)


    if not os.path.isfile(file_path):
        l_mess = f"ERROR ! File not exists: {file_path}"
        partition_messages.append(l_mess)
        log.error(l_mess)
        log.error("---")
        return partition_messages

    wb = load_workbook(path)
    l_mess = f"Загружен Excel file: {file_path}"
    log.info(l_mess)
    partition_messages.append(l_mess)

    sheet = wb.active

    with cx_Oracle.connect(user='pdd_testing', password='zA5yKk5w9nLjoSRg',
                       dsn="10.51.203.168/pdd",
                       encoding="UTF-8") as connection:
        cursor = connection.cursor()
        cursor.callproc("pdd_testing.load_questions.clean", [id_task, theme_number, partition_number, subpartition_number])
        l_mess = f"Очистили вопросы: id_task: {id_task}, theme_number: {theme_number}, " \
                 f"partition_number: {partition_number}, subpartition_number: {subpartition_number}"
        log.info(l_mess)
        partition_messages.append(l_mess)

        # Создадим новое задание
        # file_split = os.path.splitext(file_name)
        # id_theme = cursor.callfunc('admin.theme_new', int, (id_task, file_split[0]))
        # if not id_theme:
        #     print('Ошибка регистрации нового задания...')
        id_quest = 0
        id_prev_quest = -1
        order_num = 0
        id_question = 0
        for i in range(2, sheet.max_row+1):
            id_curr_quest = sheet.cell(row=i, column=1).value
            quest = sheet.cell(row=i, column=2).value
            correctly = sheet.cell(row=i, column=3).value
            answer = sheet.cell(row=i, column=4).value
            url_image = sheet.cell(row=i, column=5).value
            order_num = order_num + 1
            if not id_curr_quest:
                l_mess = f"{file_name}: WARNING ! FINISH. Преждевременное завершение загрузки. " \
                         f"Последняя загруженная строка: {id_prev_quest}"
                partition_messages.append(l_mess)
                log.warning(l_mess)
                break
            # if id_curr_quest != id_prev_quest:
            try:
                if quest:
                    id_quest = id_quest + 1
                    order_num = 1
                    id_question = cursor.callfunc("pdd_testing.load_questions.add_question", str,
                                                  [id_task, theme_number,
                                                   partition_number, subpartition_number, id_quest, f'{IMAGE_VERSION}/{url_image}', quest])

                if int(id_question) > 0:
                    cursor.callproc("pdd_testing.load_questions.add_answer", [id_question, order_num, correctly, answer])
                id_prev_quest = id_curr_quest
            except BaseException as e:
                l_mess = f"{file_name}: ERROR ! Empty id_question, Номер вопроса: {order_num}, error: {e}"
                partition_messages.append(l_mess)
                log.error(l_mess)
                break

        connection.commit()
        now = datetime.datetime.now()
        l_mess = f"Загрузка вопросов завершена: {file_name}. " \
                 f"Последняя загруженная строка: {id_prev_quest} : {now.strftime('%d-%m-%Y %H:%M:%S')}"
        log.info(l_mess)
        log.info("--")
        partition_messages.append(l_mess)
        return partition_messages


def load_task(id_task, lang):
    
    load_messages = []
    part = 1
    for sub_part in range(1, 15):
        f_name = f'ПДД/{sub_part}.xlsx'
#        f_name = f'ПДД/{sub_part}. Раздел.xlsx'
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

    f_name = 'ОБД.xlsx'
    part = 2
    load_messages.append(load_partition_questions(id_task, lang, part, 0, f'ОБД/1. {f_name}'))

    f_name = 'МЕД.xlsx'
    part = 3
    load_messages.append(load_partition_questions(id_task, lang, part, 0, f'МЕД/1. {f_name}'))

    #Административная ответственность
    f_name = 'АДМ.xlsx'
    part = 4
    load_messages.append(load_partition_questions(id_task, lang, part, 0, f'АДМ/1. {f_name}'))

    #Группа СПДД
    f_name = 'СППДД A1, A, B1.xlsx'
    part = 5
    sub_part = 1
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СППДД/1. {f_name}'))

    f_name = 'СППДД B, BE.xlsx'
    part = 5
    sub_part = 2
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СППДД/2. {f_name}'))

    f_name = 'СППДД C1, C.xlsx'
    part = 5
    sub_part = 3
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СППДД/3. {f_name}'))

    f_name = 'СППДД D1, D, Tb.xlsx'
    part = 5
    sub_part = 4
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СППДД/4. {f_name}'))

    f_name = 'СППДД C1E, CE, D1E, DE.xlsx'
    part = 5
    sub_part = 5
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СППДД/5. {f_name}'))

    f_name = 'СППДД Tm.xlsx'
    part = 5
    sub_part = 6
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СППДД/6. {f_name}'))

    #Группа СПОБД
    f_name = 'СПОБД A1, A, B1.xlsx'
    part = 6
    sub_part = 1
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СПОБД/1. {f_name}'))

    f_name = 'СПОБД C1, C.xlsx'
    part = 6
    sub_part = 3
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СПОБД/2. {f_name}'))

    f_name = 'СПОБД D1, D, Tb.xlsx'
    part = 6
    sub_part = 4
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СПОБД/3. {f_name}'))

    f_name = 'СПОБД C1E, CE, D1E, DE.xlsx'
    part = 6
    sub_part = 5
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СПОБД/4. {f_name}'))

    f_name = 'СПОБД Tm.xlsx'
    part = 6
    sub_part = 6
    load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f'СПОБД/5. {f_name}'))
    return load_messages


if __name__ == "__main__":
    if LOAD_TEST_RU:
        messages = load_task(1, 'ru')
        log.info(f"--------------> Загружены вопросы на русском языке! {messages}")
    if LOAD_TEST_KZ:
        messages = load_task(2, 'kz')
        log.info(f"--------------> Загружены вопросы на казахском языке! {messages}")
    if LOAD_TEST_EN:
        load_task(3, 'en')
        log.info(f"--------------> Загружены вопросы на английском языке! {messages}")
